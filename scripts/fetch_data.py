#!/usr/bin/env python3
"""Fetch the Bulgarian organic-production operators registry and dump static
artifacts (JSON + XLSX + meta) that the GitHub Pages site consumes.

The source API (https://bioreg.mzh.government.bg) is a jQuery DataTables
server-side endpoint. It requires a warm-up GET (for the session cookie) and an
``X-Requested-With`` header, and it returns no CORS headers — so it cannot be
called from the browser. We fetch it here, out-of-band, and commit the result.

Usage:
    python scripts/fetch_data.py [--out public/data]
"""
from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import re
import sys
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE = "https://bioreg.mzh.government.bg"
LIST_PAGE = f"{BASE}/Home/DataBaseList"

DATASETS = {
    "effective": f"{BASE}/Home/DataBaseListEffective",
    "noneffective": f"{BASE}/Home/DataBaseListNonEffective",
}

# (data key, server column name) in display order. `name` is PascalCase as the
# server expects it; `data` is the camelCase key in the JSON response.
COLUMNS = [
    ("contractCode", "ContractCode"),
    ("companyName", "CompanyName"),
    ("controllerName", "ControllerName"),
    ("districtName", "DistrictName"),
    ("activitiesText", "ActivitiesText"),
    ("certificateText", "CertificateText"),
]

# Bulgarian headers, matching the source site, used for the Excel export.
HEADERS_BG = {
    "contractCode": "Договор",
    "companyName": "Име",
    "controllerName": "Контролиращо лице",
    "districtName": "Област",
    "activitiesText": "Дейности",
    "certificateNumbers": "Сертификати",
    "certificateUrl": "Връзка към сертификат",
}

TIMEOUT = 90
HEADERS = {
    "X-Requested-With": "XMLHttpRequest",
    "Referer": LIST_PAGE,
    "User-Agent": "Mozilla/5.0 (compatible; bioreg-mirror/1.0)",
    "Accept": "application/json, text/javascript, */*; q=0.01",
}

# One certificate entry: <a href="...">NUMBER</a>, possibly relative.
_ANCHOR_RE = re.compile(
    r'<a\b[^>]*\bhref="([^"]*)"[^>]*>(.*?)</a>', re.IGNORECASE | re.DOTALL
)
_TAG_RE = re.compile(r"<[^>]+>")
# Only certificate-detail links on the known host are ever emitted.
_ALLOWED_PREFIX = f"{BASE}/Home/CertificateDetails/"


def build_payload() -> dict:
    """DataTables server-side params requesting the full dataset (length=-1)."""
    payload = {
        "draw": "1",
        "start": "0",
        "length": "-1",
        "search[value]": "",
        "search[regex]": "false",
        "order[0][column]": "0",
        "order[0][dir]": "asc",
    }
    for i, (data, name) in enumerate(COLUMNS):
        payload[f"columns[{i}][data]"] = data
        payload[f"columns[{i}][name]"] = name
        payload[f"columns[{i}][searchable]"] = "true"
        payload[f"columns[{i}][orderable]"] = "true"
        payload[f"columns[{i}][search][value]"] = ""
        payload[f"columns[{i}][search][regex]"] = "false"
    return payload


def parse_certificates(cert_html: str) -> list[dict]:
    """Turn the source's anchor markup into a validated [{number, url}] list.

    No HTML is ever forwarded to the site: each link is rebuilt to an absolute
    URL on the known host and rejected otherwise, so the front end can render it
    without trusting upstream markup.
    """
    if not cert_html:
        return []
    out = []
    for href, label in _ANCHOR_RE.findall(cert_html):
        url = href if href.startswith("http") else f"{BASE}{href}"
        if not url.startswith(_ALLOWED_PREFIX):
            continue
        number = html.unescape(_TAG_RE.sub("", label)).strip()
        if number:
            out.append({"number": number, "url": url})
    return out


def fetch(session: requests.Session, url: str) -> list[dict]:
    resp = session.post(url, data=build_payload(), headers=HEADERS, timeout=TIMEOUT)
    resp.raise_for_status()
    body = resp.json()
    if body.get("error"):
        raise RuntimeError(f"API returned error for {url}: {body['error']}")
    rows = body.get("data") or []
    total = body.get("recordsTotal", 0)
    # Guard: never let a bad/empty scrape overwrite good data.
    if total == 0 or len(rows) == 0:
        raise RuntimeError(f"Refusing empty dataset from {url} (recordsTotal={total})")
    if len(rows) != total:
        print(
            f"  warning: {url} returned {len(rows)} rows but recordsTotal={total}",
            file=sys.stderr,
        )
    return rows


def normalize(rows: list[dict]) -> list[dict]:
    out = []
    for r in rows:
        certs = parse_certificates(r.get("certificateText") or "")
        out.append(
            {
                "contractCode": (r.get("contractCode") or "").strip(),
                "companyName": (r.get("companyName") or "").strip(),
                "controllerName": (r.get("controllerName") or "").strip(),
                "districtName": (r.get("districtName") or "").strip(),
                "activitiesText": (r.get("activitiesText") or "").strip(),
                "certificates": certs,
                "certificateNumbers": ", ".join(c["number"] for c in certs),
                "certificateUrl": certs[0]["url"] if certs else "",
            }
        )
    return out


def write_json(path: Path, rows: list[dict]) -> None:
    path.write_text(
        json.dumps(rows, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )


def write_xlsx(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Регистър"
    fields = [
        "contractCode",
        "companyName",
        "controllerName",
        "districtName",
        "activitiesText",
        "certificateNumbers",
        "certificateUrl",
    ]
    ws.append([HEADERS_BG[f] for f in fields])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    url_col = fields.index("certificateUrl") + 1
    for row in rows:
        ws.append([row.get(f, "") for f in fields])
        url = row.get("certificateUrl")
        if url:
            cell = ws.cell(row=ws.max_row, column=url_col)
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")

    # Reasonable column widths.
    widths = [22, 30, 30, 16, 36, 28, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{ws.max_row}"
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--out",
        default=str(Path(__file__).resolve().parent.parent / "public" / "data"),
        help="output directory (default: public/data)",
    )
    args = ap.parse_args()
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    print(f"Warming session at {LIST_PAGE} ...")
    session.get(LIST_PAGE, headers={"User-Agent": HEADERS["User-Agent"]}, timeout=TIMEOUT)

    counts = {}
    for key, url in DATASETS.items():
        print(f"Fetching {key} from {url} ...")
        rows = normalize(fetch(session, url))
        counts[key] = len(rows)
        write_json(out_dir / f"{key}.json", rows)
        write_xlsx(out_dir / f"bioreg-{key}.xlsx", rows)
        print(f"  {key}: {len(rows)} rows -> {key}.json, bioreg-{key}.xlsx")

    meta = {
        "effectiveCount": counts.get("effective", 0),
        "nonEffectiveCount": counts.get("noneffective", 0),
        "updatedAt": dt.datetime.now(dt.timezone.utc)
        .replace(microsecond=0)
        .isoformat(),
        "source": LIST_PAGE,
    }
    (out_dir / "meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Wrote meta.json: {meta}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
